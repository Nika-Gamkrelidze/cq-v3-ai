#!/usr/bin/env python3
"""Generate the Georgian customer intake form (.xlsx).

Some customers cannot assign anyone to configure their tenant. They fill this in with
knowledge they already have, email it back, and we provision from it.

Two facts from the backend drive the design, and neither is negotiable:

* Scoring dimension `key` is auto-derived from `name` (services/scoring.py::_slug), and
  Georgian names slug cleanly. So the form asks for name/weight/guidance and NEVER a key.
  Weights must total exactly 100 -- unless every weight is blank, in which case
  scoring_store distributes them evenly. That blank-means-even rule is the escape hatch
  for a customer who cannot rank priorities, so the sheet says so out loud.
* The KB CSV importer is header-driven and turns EACH ROW into exactly one retrievable
  chunk, rendered as "header: value" lines (services/kb_ingest.py::csv_to_chunks). The
  Georgian headers therefore become part of the embedded text, and one-question-per-row
  retrieves far better than a long policy that gets sliced at CHUNK_SIZE=1000 chars.

Runs on the host's python3.9 with openpyxl only.
"""
import os

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

OUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "dist")
OUT_NAME = "CommuniQ-AI-კითხვარი.xlsx"

# ---- palette: matches the product's dark-red brand without being unreadable in print
INK = "FF1F2937"
BRAND = "FFC02637"
HEAD_BG = "FF1F2937"
HEAD_FG = "FFFFFFFF"
EXAMPLE_BG = "FFFFF7E6"
NOTE_BG = "FFF3F4F6"
OK_BG = "FFD1FAE5"
BAD_BG = "FFFEE2E2"

F_TITLE = Font(name="Sylfaen", size=16, bold=True, color=BRAND)
F_H2 = Font(name="Sylfaen", size=12, bold=True, color=INK)
F_HEAD = Font(name="Sylfaen", size=11, bold=True, color=HEAD_FG)
F_BODY = Font(name="Sylfaen", size=11, color=INK)
F_NOTE = Font(name="Sylfaen", size=10, italic=True, color="FF6B7280")
F_EXAMPLE = Font(name="Sylfaen", size=11, italic=True, color="FF92400E")

WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_C = Alignment(wrap_text=True, vertical="center", horizontal="center")
THIN = Side(style="thin", color="FFD1D5DB")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_row(ws, row, headers, widths):
    for i, (text, width) in enumerate(zip(headers, widths), start=1):
        c = ws.cell(row=row, column=i, value=text)
        c.font = F_HEAD
        c.fill = PatternFill("solid", fgColor=HEAD_BG)
        c.alignment = WRAP_C
        c.border = BOX
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.row_dimensions[row].height = 30


def _note(ws, row, text, span):
    """A full-width explanatory band. These carry the rules, so they must be readable."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = F_NOTE
    c.fill = PatternFill("solid", fgColor=NOTE_BG)
    c.alignment = WRAP
    return c


# ---------------------------------------------------------------- 1. instructions
def sheet_intro(wb):
    ws = wb.create_sheet("ინსტრუქცია")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 104

    def line(row, text, font, height=None):
        c = ws.cell(row=row, column=2, value=text)
        c.font = font
        c.alignment = WRAP
        if height:
            ws.row_dimensions[row].height = height

    line(2, "CommuniQ AI — დანერგვის კითხვარი", F_TITLE, 26)
    line(3, "შეავსეთ ეს ფაილი და დაგვიბრუნეთ. დანარჩენს ჩვენ მოვაწყობთ.", F_H2, 20)

    line(
        5,
        "ეს კითხვარი გვჭირდება იმისთვის, რომ სისტემა თქვენს წესებზე მოვარგოთ: რის მიხედვით "
        "შეფასდეს ოპერატორების ზარები და რა ინფორმაციაზე დაყრდნობით შემოწმდეს, სწორ პასუხს "
        "აძლევს თუ არა ოპერატორი კლიენტს.",
        F_BODY,
        46,
    )
    line(
        6,
        "ტექნიკური ცოდნა არ არის საჭირო. საკმარისია ის, რაც თქვენს გუნდმა ისედაც იცის.",
        F_BODY,
        18,
    )

    line(8, "რა უნდა შეავსოთ", F_H2, 20)
    steps = [
        ("1.  ფურცელი „კომპანია“", "მოკლე ინფორმაცია თქვენ შესახებ. 2 წუთი."),
        ("2.  ფურცელი „შეფასების კრიტერიუმები“",
         "რის მიხედვით უნდა შეფასდეს ზარი. უკვე შევსებულია სტანდარტული ვარიანტით — "
         "შეგიძლიათ დატოვოთ როგორც არის, შეასწოროთ ან სრულად ჩაანაცვლოთ."),
        ("3.  ფურცელი „ცოდნის ბაზა“",
         "თქვენი ხშირად დასმული კითხვები და პასუხები. ეს ყველაზე მნიშვნელოვანი ნაწილია — "
         "სწორედ ამას ადარებს სისტემა ოპერატორის ნათქვამს."),
    ]
    r = 9
    for title, body in steps:
        c = ws.cell(row=r, column=2, value=title)
        c.font = Font(name="Sylfaen", size=11, bold=True, color=INK)
        c.alignment = WRAP
        r += 1
        c = ws.cell(row=r, column=2, value="     " + body)
        c.font = F_BODY
        c.alignment = WRAP
        ws.row_dimensions[r].height = 32
        r += 1

    line(r + 1, "სასარგებლო ცოდნა", F_H2, 20)
    tips = [
        "ყვითელი ფონის სტრიქონები ნიმუშებია — შეცვალეთ ან წაშალეთ.",
        "ვარსკვლავით (*) აღნიშნული ველები აუცილებელია. დანარჩენი — სურვილისამებრ.",
        "თუ რომელიმე კითხვაზე პასუხი არ იცით, დატოვეთ ცარიელი. ჩვენ დაგიკავშირდებით.",
        "ფაილს ნუ გადაარქმევთ სახელს და ნუ წაშლით ფურცლებს — ასე უფრო სწრაფად დავამუშავებთ.",
    ]
    r += 2
    for t in tips:
        c = ws.cell(row=r, column=2, value="•  " + t)
        c.font = F_BODY
        c.alignment = WRAP
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    c = ws.cell(row=r, column=2,
                value="შევსების შემდეგ დააბრუნეთ ფაილი იმავე მისამართზე, საიდანაც მიიღეთ. "
                      "შევამოწმებთ და დაგიკავშირდებით.")
    c.font = Font(name="Sylfaen", size=11, bold=True, color=BRAND)
    c.alignment = WRAP
    ws.row_dimensions[r].height = 32
    return ws


# ---------------------------------------------------------------- 2. company
def sheet_company(wb):
    ws = wb.create_sheet("კომპანია")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 58

    c = ws.cell(row=1, column=1, value="კომპანიის შესახებ")
    c.font = F_TITLE
    ws.row_dimensions[1].height = 24

    fields = [
        ("კომპანიის სახელი *", "შპს „მაგალითი“", None),
        ("საქმიანობის სფერო *", "ბანკი", "ბანკი,სადაზღვევო,კლინიკა,ტელეკომი,სასტუმრო,ვაჭრობა,სხვა"),
        ("ქალაქი / რეგიონი", "თბილისი", None),
        ("საკონტაქტო პირი *", "ნინო მაისურაძე", None),
        ("ელფოსტა *", "nino@example.ge", None),
        ("ტელეფონი", "+995 555 00 00 00", None),
        ("ზარების ენა *", "ქართული", "ქართული,ქართული და რუსული,ქართული და ინგლისური,სამივე"),
        ("ოპერატორების რაოდენობა", "12", None),
    ]

    row = 3
    for label, example, choices in fields:
        lc = ws.cell(row=row, column=1, value=label)
        lc.font = Font(name="Sylfaen", size=11, bold=True, color=INK)
        lc.alignment = Alignment(vertical="center")
        lc.border = BOX

        vc = ws.cell(row=row, column=2, value=example)
        vc.font = F_EXAMPLE
        vc.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
        vc.alignment = Alignment(vertical="center")
        vc.border = BOX
        ws.row_dimensions[row].height = 22

        if choices:
            dv = DataValidation(type="list", formula1='"%s"' % choices, allow_blank=True)
            dv.error = "აირჩიეთ სიიდან."
            dv.errorTitle = "არასწორი მნიშვნელობა"
            ws.add_data_validation(dv)
            dv.add(vc)
        row += 1

    _note(ws, row + 1,
          "ყვითელ უჯრებში ჩაწერილი ტექსტი ნიმუშია — ჩაანაცვლეთ თქვენი მონაცემებით.", 2)
    ws.row_dimensions[row + 1].height = 20
    return ws


# ---------------------------------------------------------------- 3. rubric
# Pre-filled with a complete, working rubric that already totals 100. A customer with no
# time can keep it verbatim and still be onboarded correctly -- which is the entire point
# of this form. Every guidance string describes OBSERVABLE behaviour, because guidance is
# injected verbatim into the scoring prompt and is what separates a 90 from a 40.
RUBRIC = [
    ("მისალმება და იდენტიფიკაცია", 10,
     "მიესალმა, დაასახელა კომპანია და თავისი სახელი, დაადასტურა კლიენტის ვინაობა "
     "დადგენილი წესით."),
    ("მოსმენა და ემპათია", 15,
     "არ შეაწყვეტინა კლიენტს, საკუთარი სიტყვებით გაიმეორა პრობლემა დასადასტურებლად, "
     "გამოხატა თანაგრძნობა."),
    ("პრობლემის სწორად დადგენა", 20,
     "დასვა დამაზუსტებელი კითხვები და სწორად დაადგინა მიზეზი სხვა თანამშრომელთან "
     "ზედმეტი გადამისამართების გარეშე."),
    ("ინფორმაციის სიზუსტე", 25,
     "მიაწოდა ზუსტი ინფორმაცია ტარიფებზე, ვადებსა და პირობებზე; არ დაპირდა იმას, რაც "
     "კომპანიის წესებით გათვალისწინებული არ არის."),
    ("საკითხის მოგვარება", 20,
     "შესთავაზა კონკრეტული გადაწყვეტა ან შემდეგი ნაბიჯი მკაფიო ვადით და დაადასტურა, "
     "რომ პასუხმა კლიენტი დააკმაყოფილა."),
    ("ზრდილობა და ტონი", 10,
     "მთელი საუბრის განმავლობაში ისაუბრა თავაზიანად და მშვიდად, მათ შორის მაშინ, როცა "
     "კლიენტი უკმაყოფილო იყო."),
]
RUBRIC_ROWS = 24  # room to add their own without touching the file's structure


def sheet_rubric(wb):
    ws = wb.create_sheet("შეფასების კრიტერიუმები")
    ws.sheet_view.showGridLines = False

    c = ws.cell(row=1, column=1, value="რის მიხედვით შეფასდეს ზარი")
    c.font = F_TITLE
    ws.row_dimensions[1].height = 24

    _note(ws, 2,
          "ქვემოთ უკვე შევსებულია სტანდარტული შეფასების სისტემა. თუ ის თქვენთვის მისაღებია — "
          "არაფერი შეცვალოთ. სურვილისამებრ შეასწორეთ ტექსტი, წონები, ან დაამატეთ თქვენი "
          "კრიტერიუმები ცარიელ სტრიქონებში.", 3)
    ws.row_dimensions[2].height = 44

    _note(ws, 3,
          "წონების ჯამი უნდა იყოს ზუსტად 100. თუ არ იცით რომელი კრიტერიუმი რამდენად "
          "მნიშვნელოვანია — წაშალეთ ყველა წონა და დატოვეთ სვეტი სრულიად ცარიელი; მაშინ "
          "ყველა კრიტერიუმს თანაბარ მნიშვნელობას მივანიჭებთ.", 3)
    ws.row_dimensions[3].height = 44

    head = 5
    _header_row(ws, head,
                ["კრიტერიუმი", "წონა (%)", "მითითება — რას აკეთებს კარგი ოპერატორი"],
                [30, 11, 82])

    first = head + 1
    for i in range(RUBRIC_ROWS):
        r = first + i
        example = RUBRIC[i] if i < len(RUBRIC) else None
        for col in (1, 2, 3):
            cell = ws.cell(row=r, column=col)
            cell.border = BOX
            cell.alignment = WRAP if col != 2 else WRAP_C
            if example:
                cell.value = example[col - 1]
                cell.font = F_EXAMPLE
                cell.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
            else:
                cell.font = F_BODY
        ws.row_dimensions[r].height = 34 if example else 22

    last = first + RUBRIC_ROWS - 1
    total = last + 1
    tc = ws.cell(row=total, column=1, value="ჯამი")
    tc.font = Font(name="Sylfaen", size=11, bold=True, color=INK)
    tc.alignment = Alignment(horizontal="right", vertical="center")
    tc.border = BOX

    sc = ws.cell(row=total, column=2, value="=SUM(B%d:B%d)" % (first, last))
    sc.font = Font(name="Sylfaen", size=12, bold=True, color=INK)
    sc.alignment = WRAP_C
    sc.border = BOX

    # The one error that would otherwise bounce the form back over email. Caught live,
    # at the customer's desk, while they are still looking at the numbers.
    ref = "B%d" % total
    ws.conditional_formatting.add(ref, CellIsRule(
        operator="equal", formula=["100"],
        fill=PatternFill("solid", start_color=OK_BG, end_color=OK_BG)))
    ws.conditional_formatting.add(ref, CellIsRule(
        operator="notEqual", formula=["100"],
        fill=PatternFill("solid", start_color=BAD_BG, end_color=BAD_BG)))

    hc = ws.cell(row=total, column=3,
                 value="მწვანე = სწორია. წითელი = ჯამი 100 არ არის, შეასწორეთ წონები "
                       "(ან დატოვეთ სვეტი მთლიანად ცარიელი).")
    hc.font = F_NOTE
    hc.alignment = WRAP
    hc.border = BOX
    ws.row_dimensions[total].height = 30

    _note(ws, total + 2,
          "მითითება ყველაზე მნიშვნელოვანი ველია: ზუსტად ის ტექსტი ეუბნება სისტემას, როგორ "
          "განასხვაოს კარგი ზარი საშუალოსგან. დაწერეთ ის, რაც ჩანაწერში შესამჩნევია.", 3)
    ws.row_dimensions[total + 2].height = 32

    good = ws.cell(row=total + 3, column=1, value="სუსტი მითითება:")
    good.font = Font(name="Sylfaen", size=10, bold=True, color="FF991B1B")
    c = ws.cell(row=total + 3, column=3, value="„იყოს თავაზიანი და პროფესიონალი.“")
    c.font = F_NOTE
    c.alignment = WRAP

    good = ws.cell(row=total + 4, column=1, value="ძლიერი მითითება:")
    good.font = Font(name="Sylfaen", size=10, bold=True, color="FF065F46")
    c = ws.cell(row=total + 4, column=3,
                value="„მიესალმა სახელით, არ შეაწყვეტინა, გაიმეორა პრობლემა "
                      "დასადასტურებლად და დაასახელა კონკრეტული ვადა.“")
    c.font = F_NOTE
    c.alignment = WRAP
    ws.row_dimensions[total + 4].height = 30
    return ws


# ---------------------------------------------------------------- 4. knowledge base
# Column headers are load-bearing: csv_to_chunks renders each row as "header: value"
# lines, so these Georgian words end up inside the embedded text and shape retrieval.
KB_HEADERS = ["კითხვა", "პასუხი", "კატეგორია", "ხილვადობა"]
KB_EXAMPLES = [
    ("რამდენ ხანში მზადდება ბარათი?",
     "სტანდარტული ბარათი მზადდება 3 სამუშაო დღეში და გაიცემა იმ ფილიალში, სადაც განაცხადი "
     "შეიტანეთ. სასწრაფო გაცემა შესაძლებელია 1 სამუშაო დღეში, დამატებითი საკომისიოთი.",
     "ბარათები", "საჯარო"),
    ("როგორ დავბლოკო დაკარგული ბარათი?",
     "დარეკეთ ცხელ ხაზზე 24 საათის განმავლობაში ან დაბლოკეთ ბარათი მობილურ აპლიკაციაში "
     "განყოფილებიდან „ბარათები“. ბლოკირება მყისიერად ხდება და საკომისიოს არ საჭიროებს.",
     "უსაფრთხოება", "საჯარო"),
    ("რა ღირს ანგარიშის თვიური მომსახურება?",
     "ფიზიკური პირის სტანდარტული ანგარიშის მომსახურება ღირს 5 ლარი თვეში. ხელფასის "
     "პროექტის მონაწილეებისთვის მომსახურება უფასოა.",
     "ტარიფები", "საჯარო"),
    ("რა ვადაში უნდა განიხილოს ოპერატორმა პრეტენზია?",
     "პრეტენზია რეგისტრირდება მიღებისთანავე და განიხილება 10 სამუშაო დღეში. თუ საკითხი "
     "საერთაშორისო ტრანზაქციას ეხება, ვადა 30 დღემდე გრძელდება.",
     "პრეტენზიები", "საჯარო"),
    ("როდის უნდა გადავამისამართო ზარი უფროსთან?",
     "გადაამისამართეთ, თუ კლიენტი ითხოვს საკომისიოს ჩამოწერას 50 ლარზე მეტი ოდენობით, "
     "ან თუ საუბარი კონფლიქტურ ფაზაში გადადის. ოპერატორს დამოუკიდებლად შეუძლია ჩამოწეროს "
     "50 ლარამდე თანხა თვეში ერთხელ.",
     "შიდა წესები", "შიდა"),
]
KB_ROWS = 60


def sheet_kb(wb):
    ws = wb.create_sheet("ცოდნის ბაზა")
    ws.sheet_view.showGridLines = False

    c = ws.cell(row=1, column=1, value="თქვენი კითხვები და პასუხები")
    c.font = F_TITLE
    ws.row_dimensions[1].height = 24

    _note(ws, 2,
          "ჩაწერეთ ის კითხვები, რომლებსაც კლიენტები ყველაზე ხშირად სვამენ, და სწორი პასუხები. "
          "სწორედ ამას შეადარებს სისტემა ოპერატორის ნათქვამს და დაგიფიქსირებთ, თუ ოპერატორმა "
          "არასწორი ინფორმაცია მიაწოდა.", 4)
    ws.row_dimensions[2].height = 44

    _note(ws, 3,
          "მთავარი წესი: ერთი სტრიქონი — ერთი კითხვა და ერთი დასრულებული პასუხი. ნუ ჩასვამთ "
          "მთელ დოკუმენტს ერთ უჯრაში. სჯობს 20 მოკლე სტრიქონი, ვიდრე 3 გრძელი. პასუხი "
          "დამოუკიდებლად უნდა იკითხებოდეს — არ მიუთითოთ „იხილეთ ზემოთ“.", 4)
    ws.row_dimensions[3].height = 44

    _note(ws, 4,
          "ხილვადობა: „საჯარო“ — პასუხი შეიძლება სიტყვასიტყვით ეთქვას კლიენტს. "
          "„შიდა“ — მხოლოდ თანამშრომლებისთვის (შიდა წესები, ფასდაკლების ლიმიტები, "
          "ესკალაციის პირობები). თუ ეჭვობთ — აირჩიეთ „შიდა“.", 4)
    ws.row_dimensions[4].height = 44

    head = 6
    _header_row(ws, head, KB_HEADERS, [42, 78, 18, 14])

    dv = DataValidation(type="list", formula1='"საჯარო,შიდა"', allow_blank=True)
    dv.error = "აირჩიეთ „საჯარო“ ან „შიდა“."
    dv.errorTitle = "არასწორი მნიშვნელობა"
    ws.add_data_validation(dv)

    first = head + 1
    for i in range(KB_ROWS):
        r = first + i
        example = KB_EXAMPLES[i] if i < len(KB_EXAMPLES) else None
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.border = BOX
            cell.alignment = WRAP_C if col == 4 else WRAP
            if example:
                cell.value = example[col - 1]
                cell.font = F_EXAMPLE
                cell.fill = PatternFill("solid", fgColor=EXAMPLE_BG)
            else:
                cell.font = F_BODY
        dv.add(ws.cell(row=r, column=4))
        ws.row_dimensions[r].height = 46 if example else 30

    ws.freeze_panes = ws.cell(row=first, column=1)
    return ws


def main():
    wb = Workbook()
    wb.remove(wb.active)

    sheet_intro(wb)
    sheet_company(wb)
    sheet_rubric(wb)
    sheet_kb(wb)

    wb.properties.title = "CommuniQ AI — დანერგვის კითხვარი"
    wb.properties.subject = "ტენანტის კონფიგურაცია"
    wb.properties.creator = "CommuniQ"

    if not os.path.isdir(OUT_DIR):
        os.makedirs(OUT_DIR)
    path = os.path.join(OUT_DIR, OUT_NAME)
    wb.save(path)
    print("wrote %s" % path)
    return path


if __name__ == "__main__":
    main()
